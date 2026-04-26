"""Lee SUNAT exportaciones.xlsx y devuelve un DataFrame tidy en formato largo."""
import os
from pathlib import Path
import pandas as pd
import openpyxl
from typing import Optional

PROJECT_ROOT = Path(__file__).resolve().parent.parent

MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Setiembre", "Octubre", "Noviembre", "Diciembre"]
MES_NUM = {m: i + 1 for i, m in enumerate(MESES)}


def _resolve_xlsx_path(path: Optional[Path] = None) -> Path:
    if path is not None:
        return Path(path)

    env_path = os.getenv("SUNAT_XLSX_PATH")
    candidates = [
        Path(env_path) if env_path else None,
        PROJECT_ROOT / "sunat" / "data" / "sunat.xlsx",
        PROJECT_ROOT / "SUNAT exportaciones.xlsx",
        PROJECT_ROOT / "sunat.xlsx",
        PROJECT_ROOT / "sunat_x2" / "sunat.xlsx",
    ]
    valid_candidates = [p for p in candidates if p is not None]
    for candidate in valid_candidates:
        if candidate.exists():
            return candidate

    searched = "\n".join(f"- {p}" for p in valid_candidates)
    raise FileNotFoundError(
        "No se encontro el archivo SUNAT.\n"
        "Colocalo en una de estas rutas o define la variable SUNAT_XLSX_PATH:\n"
        f"{searched}"
    )


def _detect_months(ws) -> list[tuple[int, int]]:
    """Devuelve [(col_idx, mes_num), ...] leyendo la fila de cabecera."""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        cols = []
        for idx, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip() in MES_NUM:
                cols.append((idx, MES_NUM[cell.strip()]))
        if cols:
            return cols
    return []


def _parse_sheet(ws, year: int) -> list[dict]:
    rows = []
    categoria = subsector = None
    month_cols = _detect_months(ws)
    if not month_cols:
        return rows
    for row in ws.iter_rows(min_row=1, values_only=True):
        # Columnas: B=1 nota, C=2 categoría, D=3 subsector, E=4 producto
        if len(row) < 5:
            continue
        c_cat, c_sub, c_prod = row[2], row[3], row[4]

        # ignorar filas de notas o totalmente vacías
        if c_cat and isinstance(c_cat, str) and c_cat.strip().lower().startswith(("nota", "fuente", "elaborac")):
            break

        if c_cat and not c_sub and not c_prod:
            cat_str = str(c_cat).strip()
            if cat_str.lower() == "total general":
                continue  # lo recalcula el dashboard
            categoria = cat_str
            subsector = None
            # Sólo "Otros" se registra desde la fila-categoría (no tiene desglose por producto)
            is_leaf = cat_str.lower().startswith("otros")
            if is_leaf and any(row[i] is not None for i, _ in month_cols if i < len(row)):
                for col_idx, mes_num in month_cols:
                    if col_idx >= len(row):
                        continue
                    valor = row[col_idx]
                    if valor is None:
                        continue
                    try:
                        v = float(valor)
                    except (TypeError, ValueError):
                        continue
                    rows.append({
                        "anio": year,
                        "mes": mes_num,
                        "fecha": pd.Timestamp(year=year, month=mes_num, day=1),
                        "categoria": cat_str,
                        "subsector": cat_str,
                        "producto": cat_str,
                        "valor_usd": v,
                    })
            continue
        if c_sub and not c_prod:
            subsector = str(c_sub).strip()
            continue
        if c_prod:
            producto = str(c_prod).strip()
            for col_idx, mes_num in month_cols:
                if col_idx >= len(row):
                    continue
                valor = row[col_idx]
                if valor is None:
                    continue
                try:
                    v = float(valor)
                except (TypeError, ValueError):
                    continue
                rows.append({
                    "anio": year,
                    "mes": mes_num,
                    "fecha": pd.Timestamp(year=year, month=mes_num, day=1),
                    "categoria": categoria or "",
                    "subsector": subsector or producto,
                    "producto": producto,
                    "valor_usd": v,
                })
    return rows


def load_long(path: Optional[Path] = None) -> pd.DataFrame:
    path = _resolve_xlsx_path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    all_rows: list[dict] = []
    for sheet_name in wb.sheetnames:
        try:
            year = int(sheet_name)
        except ValueError:
            continue
        all_rows.extend(_parse_sheet(wb[sheet_name], year))
    df = pd.DataFrame(all_rows)
    df["categoria"] = df["categoria"].replace({
        "Productos Tradicionales": "Tradicional",
        "Productos no Tradicionales": "No Tradicional",
        "Otros 5/": "Otros",
        "Otros": "Otros",
    })
    df["valor_musd"] = df["valor_usd"] / 1_000_000
    return df


if __name__ == "__main__":
    df = load_long()
    print(df.head())
    print(f"\nFilas: {len(df):,}")
    print(f"Años: {df['anio'].min()}–{df['anio'].max()}")
    print(f"Categorías: {df['categoria'].unique()}")
    print(f"Productos únicos: {df['producto'].nunique()}")
    print(f"Total acumulado: ${df['valor_usd'].sum()/1e9:,.2f} mil millones USD")
